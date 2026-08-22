"""
data_pipeline/importers/investments/b3_movimentacao.py
======================================================
Parser do arquivo de Movimentação exportado pelo portal da B3.
(investidor.b3.com.br → Extratos e Informativos → Movimentação)

Colunas esperadas:
    Entrada/Saída | Data | Movimentação | Produto | Instituição |
    Quantidade | Preço unitário | Valor da Operação

Importa:
  * Dividendos, JCP, Rendimentos de FII, Amortizações em `dividends`.
  * Bonificações, desdobros e operações sem contrapartida financeira em
    `investment_transactions` (preço/valor = 0 quando aplicável).

Ignora (com contagem):
  * Compra/venda comum (já vem do arquivo Negociação).
  * Eventos não suportados pelo schema atual.

Idempotência via `external_id` em ambas as tabelas.

Otimização (2026-05-22):
  Antes: 1 SAVEPOINT + 3-4 queries por linha do XLSX. Depois: parse 100%
  em memória + ~7 queries em batch (independente do número de linhas).
"""
from __future__ import annotations

import io
import logging
from typing import Any

import openpyxl
from sqlalchemy.engine import Engine

from core.config import settings

from .common import (
    batch_filter_existing_external_ids,
    batch_get_or_create_assets,
    batch_insert_dividends,
    batch_insert_investment_transactions,
    classify_movement,
    classify_ticker,
    ensure_external_id_columns,
    finalize_summary,
    get_or_create_b3_account,
    make_external_id,
    make_summary,
    parse_date_br,
    parse_ticker_from_produto,
    safe_error,
    to_float_br,
)

logger = logging.getLogger(__name__)

SOURCE = "b3_movimentacao"
SHEET_HINT = "movimenta"


def parse(file_bytes: bytes, engine: Engine) -> dict[str, Any]:
    """
    Processa o XLSX de Movimentação da B3 e grava proventos/eventos no app4.
    """
    summary = make_summary(SOURCE)
    user_id = settings.OWNER_USER_ID
    if not user_id:
        summary["status"] = "failed"
        summary["errors"].append("OWNER_USER_ID nao configurado.")
        return finalize_summary(summary)

    ensure_external_id_columns(engine)

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as exc:
        summary["status"] = "failed"
        summary["errors"].append(f"Arquivo invalido: {safe_error(exc)}")
        return finalize_summary(summary)

    sheet = None
    for name in wb.sheetnames:
        if SHEET_HINT in name.lower():
            sheet = wb[name]
            break
    if sheet is None:
        summary["status"] = "failed"
        summary["errors"].append(
            f"Aba 'Movimentacao' nao encontrada. Abas: {wb.sheetnames}"
        )
        return finalize_summary(summary)

    # Estratégia em 2 fases (ver b3_negociacao.py para racional).
    tx_candidates: list[dict] = []
    div_candidates: list[dict] = []

    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        if i == 0:
            continue
        try:
            parsed = _parse_row(row)
        except Exception as exc:  # noqa: BLE001
            summary["errors"].append(f"Linha {i + 1}: {safe_error(exc)}")
            continue
        if parsed is None:
            summary["rows_skipped"] += 1
            continue
        if parsed["kind"] == "income":
            div_candidates.append(parsed)
        else:
            tx_candidates.append(parsed)

    if not tx_candidates and not div_candidates:
        return finalize_summary(summary)

    try:
        with engine.connect() as conn:
            with conn.begin():
                _account_id = get_or_create_b3_account(conn, user_id)

                # 1) Resolve asset_ids (tickers de TX + de DIV)
                asset_items = []
                for c in (*tx_candidates, *div_candidates):
                    asset_items.append((
                        c["ticker"],
                        c["asset_name"],
                        classify_ticker(c["ticker"]),
                    ))
                ticker_to_id = batch_get_or_create_assets(conn, asset_items)

                # 2) Filtra duplicatas (ambas tabelas)
                tx_ext = [c["external_id"] for c in tx_candidates]
                div_ext = [c["external_id"] for c in div_candidates]
                existing_tx = batch_filter_existing_external_ids(
                    conn, "investment_transactions", tx_ext
                ) if tx_ext else set()
                existing_div = batch_filter_existing_external_ids(
                    conn, "dividends", div_ext
                ) if div_ext else set()

                # 3a) Monta linhas novas — transactions
                tx_rows = []
                for c in tx_candidates:
                    if c["external_id"] in existing_tx:
                        summary["duplicates_skipped"] += 1
                        continue
                    aid = ticker_to_id.get(c["ticker"])
                    if aid is None:
                        summary["errors"].append(
                            f"asset_id ausente para ticker={c['ticker']}"
                        )
                        continue
                    tx_rows.append({
                        "user_id":          user_id,
                        "asset_id":         aid,
                        "type":             c["tx_type"],
                        "quantity":         c["quantity"],
                        "unit_price":       c["unit_price"],
                        "fees":             0.0,
                        "transaction_date": c["transaction_date"],
                        "broker":           "B3",
                        "external_id":      c["external_id"],
                    })

                # 3b) Monta linhas novas — dividends
                div_rows = []
                for c in div_candidates:
                    if c["external_id"] in existing_div:
                        summary["duplicates_skipped"] += 1
                        continue
                    aid = ticker_to_id.get(c["ticker"])
                    if aid is None:
                        summary["errors"].append(
                            f"asset_id ausente para ticker={c['ticker']}"
                        )
                        continue
                    div_rows.append({
                        "user_id":         user_id,
                        "asset_id":        aid,
                        "type":            c["div_type"],
                        "amount_per_unit": c["amount_per_unit"],
                        "quantity":        c["quantity"],
                        "total_amount":    c["total_amount"],
                        "ex_date":         None,
                        "payment_date":    c["transaction_date"],
                        "external_id":     c["external_id"],
                    })

                # 4) Insert em lote
                inserted_tx = batch_insert_investment_transactions(conn, tx_rows)
                inserted_div = batch_insert_dividends(conn, div_rows)
                summary["transactions_imported"] += inserted_tx
                summary["incomes_imported"] += inserted_div
    except Exception as exc:  # noqa: BLE001
        summary["status"] = "failed"
        summary["errors"].append(f"Batch DB falhou: {safe_error(exc)}")

    return finalize_summary(summary)


def _parse_row(row) -> dict | None:
    """Parse puro (sem DB) de uma linha do XLSX de Movimentação.

    Retorna dict com chave "kind" in {"income", "transaction"} e payload
    correspondente, ou None se a linha deve ser ignorada.
    """
    if not row or len(row) < 8:
        return None

    entrada, data_raw, mov_raw, produto_raw, _inst, qtd_raw, preco_raw, valor_raw = row[:8]
    if not mov_raw or not produto_raw or not data_raw:
        return None

    classification = classify_movement(str(mov_raw), str(entrada or ""))
    if classification is None:
        return None
    category, canonical_type = classification
    if category == "skip":
        return None

    ticker_clean, name_from_produto = parse_ticker_from_produto(str(produto_raw))
    if not ticker_clean:
        return None

    tx_date = parse_date_br(data_raw)
    if tx_date is None:
        return None

    qtd = to_float_br(qtd_raw)
    preco = to_float_br(preco_raw)
    valor = to_float_br(valor_raw)

    ext_id = make_external_id(
        "b3mov",
        [tx_date.isoformat(), str(mov_raw).strip().lower(),
         ticker_clean, str(entrada or "").lower(), qtd_raw, valor_raw],
    )

    if category == "income":
        if valor is None or valor <= 0:
            return None
        if qtd and qtd > 0:
            apu = round(valor / qtd, 6)
            quantity_used = qtd
        else:
            apu = valor
            quantity_used = 1.0
        return {
            "kind":             "income",
            "ticker":           ticker_clean,
            "asset_name":       name_from_produto or ticker_clean,
            "div_type":         canonical_type,
            "amount_per_unit":  apu,
            "quantity":         quantity_used,
            "total_amount":     valor,
            "transaction_date": tx_date,
            "external_id":      ext_id,
        }

    # category == "transaction"
    if canonical_type not in ("buy", "sell"):
        return None
    if qtd is None or qtd <= 0:
        return None
    if (preco is None or preco == 0) and valor and valor > 0 and qtd > 0:
        preco = round(valor / qtd, 6)
    preco = preco or 0.0

    return {
        "kind":             "transaction",
        "ticker":           ticker_clean,
        "asset_name":       name_from_produto or ticker_clean,
        "tx_type":          canonical_type,
        "quantity":         qtd,
        "unit_price":       preco,
        "transaction_date": tx_date,
        "external_id":      ext_id,
    }
