"""
data_pipeline/importers/investments/b3_negociacao.py
====================================================
Parser do arquivo de Negociação exportado pelo portal da B3.
(investidor.b3.com.br → Extratos e Informativos → Negociação)

Colunas esperadas (em ordem):
    Data do Negócio | Tipo de Movimentação | Mercado | Prazo/Vencimento
    Instituição | Código de Negociação | Quantidade | Preço | Valor

Comportamento:
  * Só importa Compra/Venda (movimentações financeiras puras).
  * Cria ativos novos automaticamente em `assets`.
  * Cria instituição/conta agregadora "B3 - Carteira Consolidada".
  * Idempotente via `investment_transactions.external_id`.

Otimização (2026-05-22):
  Antes: 1 SAVEPOINT + 3-4 queries por linha do XLSX (300+ linhas = 5-10min).
  Depois: 2 fases — Parse 100% em memória, depois ~5 queries em batch
  (~300 linhas em <30s mesmo com latência alta Supabase).
"""
from __future__ import annotations

import io
import logging
from typing import Any

import openpyxl
from sqlalchemy.engine import Engine

from core.config import settings
from .common import (
    batch_filter_existing_external_ids,
    batch_get_or_create_assets,
    batch_insert_investment_transactions,
    classify_ticker,
    ensure_external_id_columns,
    finalize_summary,
    get_or_create_b3_account,
    make_external_id,
    make_summary,
    parse_date_br,
    parse_ticker_from_produto,
    safe_error,
    to_float_br,
)

logger = logging.getLogger(__name__)

SOURCE = "b3_negociacao"
SHEET_HINT = "negocia"  # encontra "Negociação", "Negociacao", etc.


def parse(file_bytes: bytes, engine: Engine) -> dict[str, Any]:
    """
    Processa o XLSX de Negociação da B3 e grava as operações no app4.

    Retorna o resumo padronizado descrito em
    `.claude/skills/investment-imports/SKILL.md`.
    """
    summary = make_summary(SOURCE)
    user_id = settings.OWNER_USER_ID
    if not user_id:
        summary["status"] = "failed"
        summary["errors"].append("OWNER_USER_ID nao configurado.")
        return finalize_summary(summary)

    ensure_external_id_columns(engine)

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as exc:
        summary["status"] = "failed"
        summary["errors"].append(f"Arquivo invalido: {safe_error(exc)}")
        return finalize_summary(summary)

    sheet = None
    for name in wb.sheetnames:
        if SHEET_HINT in name.lower():
            sheet = wb[name]
            break
    if sheet is None:
        summary["status"] = "failed"
        summary["errors"].append(
            f"Aba 'Negociacao' nao encontrada. Abas: {wb.sheetnames}"
        )
        return finalize_summary(summary)

    # Estratégia em 2 fases:
    #   FASE 1 (CPU, sem DB): parse de todas as linhas em memória.
    #   FASE 2 (DB, ~5 queries totais):
    #     - get_or_create_b3_account
    #     - batch_get_or_create_assets (1 SELECT + 1 INSERT)
    #     - batch_filter_existing_external_ids (1 SELECT — descobre duplicatas)
    #     - batch_insert_investment_transactions (1 INSERT em lote)
    #
    # Erros de parsing por linha são acumulados no summary["errors"] mas
    # nao abortam o batch. Erros de DB no batch abortam tudo (raro: o
    # batch só roda apos validação Python rigorosa).
    candidates: list[dict] = []
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        if i == 0:
            continue
        try:
            parsed = _parse_row(row, account_label="B3")
        except Exception as exc:  # noqa: BLE001
            summary["errors"].append(f"Linha {i + 1}: {safe_error(exc)}")
            continue
        if parsed is None:
            summary["rows_skipped"] += 1
            continue
        candidates.append(parsed)

    if not candidates:
        return finalize_summary(summary)

    try:
        with engine.connect() as conn:
            with conn.begin():
                # Garante setup (account agregadora)
                _account_id = get_or_create_b3_account(conn, user_id)

                # 1) Resolve asset_ids em batch
                asset_items = [
                    (c["ticker"], c["asset_name"], classify_ticker(c["ticker"]))
                    for c in candidates
                ]
                ticker_to_id = batch_get_or_create_assets(conn, asset_items)

                # 2) Descobre quais external_ids já existem (idempotência)
                all_ext_ids = [c["external_id"] for c in candidates]
                existing_ext_ids = batch_filter_existing_external_ids(
                    conn, "investment_transactions", all_ext_ids
                )

                # 3) Filtra novos e enriquece com asset_id, user_id
                rows_to_insert = []
                for c in candidates:
                    if c["external_id"] in existing_ext_ids:
                        summary["duplicates_skipped"] += 1
                        continue
                    aid = ticker_to_id.get(c["ticker"])
                    if aid is None:
                        # Não deveria acontecer (batch_get_or_create_assets cria),
                        # mas defensivo.
                        summary["errors"].append(
                            f"asset_id ausente para ticker={c['ticker']}"
                        )
                        continue
                    rows_to_insert.append({
                        "user_id":          user_id,
                        "asset_id":         aid,
                        "type":             c["tx_type"],
                        "quantity":         c["quantity"],
                        "unit_price":       c["unit_price"],
                        "fees":             0.0,
                        "transaction_date": c["transaction_date"],
                        "broker":           c["broker"],
                        "external_id":      c["external_id"],
                    })

                # 4) Insert em lote
                inserted = batch_insert_investment_transactions(conn, rows_to_insert)
                summary["transactions_imported"] += inserted
    except Exception as exc:  # noqa: BLE001
        summary["status"] = "failed"
        summary["errors"].append(f"Batch DB falhou: {safe_error(exc)}")

    return finalize_summary(summary)


def _parse_row(row, *, account_label: str) -> dict | None:
    """Parse puro (sem DB) de uma linha do XLSX de Negociação.

    Retorna dict com chaves: ticker, asset_name, tx_type, quantity,
    unit_price, transaction_date, broker, external_id.
    Retorna None se a linha deve ser ignorada (sem dados, tipo desconhecido).
    """
    if not row or len(row) < 9:
        return None

    data_raw, tipo_raw, mercado_raw, _prazo, _inst, ticker_raw, qtd_raw, preco_raw, valor_raw = row[:9]
    if not ticker_raw or not data_raw or not tipo_raw:
        return None

    tipo_norm = str(tipo_raw).strip().lower()
    if tipo_norm == "compra":
        tx_type = "buy"
    elif tipo_norm == "venda":
        tx_type = "sell"
    else:
        return None

    ticker_clean, name_from_produto = parse_ticker_from_produto(str(ticker_raw))
    if not ticker_clean:
        return None

    qtd = to_float_br(qtd_raw)
    preco = to_float_br(preco_raw)
    valor = to_float_br(valor_raw)
    if qtd is None or qtd <= 0:
        return None

    if (valor is None or valor == 0) and preco and preco > 0:
        valor = round(preco * qtd, 2)
    if (preco is None or preco == 0) and valor and valor > 0 and qtd > 0:
        preco = round(valor / qtd, 6)
    preco = preco or 0.0

    tx_date = parse_date_br(data_raw)
    if tx_date is None:
        return None

    ext_id = make_external_id(
        "b3neg",
        [tx_date.isoformat(), tx_type, ticker_clean, qtd, preco, str(mercado_raw or "")],
    )

    return {
        "ticker":            ticker_clean,
        "asset_name":        name_from_produto or ticker_clean,
        "tx_type":           tx_type,
        "quantity":          qtd,
        "unit_price":        preco,
        "transaction_date":  tx_date,
        "broker":            account_label,
        "external_id":       ext_id,
    }
