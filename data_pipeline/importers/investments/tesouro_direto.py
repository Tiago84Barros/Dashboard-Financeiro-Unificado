"""Importador do Extrato Consolidado do Tesouro Direto (.xlsx).

O arquivo e uma fotografia mensal das posicoes, portanto alimenta
``portfolio_position_snapshots``. Ele nunca cria operacoes de compra/venda.
Antes de gravar, titulos cuja identidade economica ja possui operacao importada
da B3 sao recusados: a B3 permanece a fonte canonica nesses casos.
"""
from __future__ import annotations

import calendar
import io
import re
import unicodedata
from datetime import date, datetime
from typing import Any, Iterable

import openpyxl
from sqlalchemy import text
from sqlalchemy.engine import Connection, Engine

from core.config import settings

from .common import (
    ensure_external_id_columns,
    get_or_create_asset,
    make_external_id,
    make_summary,
    finalize_summary,
    safe_error,
)
from .positions import PORTFOLIO_NAME, PORTFOLIO_TYPE
from .xp_consolidado import _ensure_portfolio, _insert_snapshot

SOURCE = "tesouro_direto"
MAX_FILE_BYTES = 5 * 1024 * 1024
_PERIOD_RE = re.compile(r"\b(0[1-9]|1[0-2])/(20\d{2})\b")


def _norm(value: Any) -> str:
    text_value = str(value or "").strip()
    return unicodedata.normalize("NFKD", text_value).encode("ascii", "ignore").decode().upper()


def _maturity_year(value: Any) -> str:
    match = re.search(r"\b(20\d{2})\b", str(value or ""))
    return match.group(1) if match else "X"


def tesouro_security_key(title: Any, maturity: Any) -> str:
    """Normaliza titulo/vencimento em identificador estavel entre fontes."""
    normalized = _norm(title)
    if re.fullmatch(r"T(?:SELIC|IPCAJ?|PREJ?|EDUCA|RENDA)20\d{2}", normalized):
        return normalized
    year = _maturity_year(maturity) or _maturity_year(normalized)
    if "SELIC" in normalized or "LFT" in normalized:
        kind = "TSELIC"
    elif "EDUCA" in normalized:
        kind = "TEDUCA"
    elif "RENDA" in normalized:
        kind = "TRENDA"
    elif "IPCA" in normalized or "NTN-B" in normalized:
        has_coupon = "JUROS" in normalized or (
            "NTN-B" in normalized and "PRINCIPAL" not in normalized
        )
        kind = "TIPCAJ" if has_coupon else "TIPCA"
    elif "PREFIXADO" in normalized or "PRE" in normalized or "LTN" in normalized or "NTN-F" in normalized:
        kind = "TPREJ" if ("JUROS" in normalized or "NTN-F" in normalized) else "TPRE"
    else:
        kind = "TESOUR"
    return f"{kind}{year}"


def _number(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    raw = str(value).replace("R$", "").replace(" ", "").strip()
    if "," in raw:
        raw = raw.replace(".", "").replace(",", ".")
    try:
        return float(raw)
    except ValueError:
        return None


def _report_date(rows: Iterable[tuple[Any, ...]]) -> date | None:
    for row in rows:
        for value in row:
            match = _PERIOD_RE.search(str(value or ""))
            if match:
                month, year = int(match.group(1)), int(match.group(2))
                return date(year, month, calendar.monthrange(year, month)[1])
    return None


def _parse_rows(rows: list[tuple[Any, ...]]) -> tuple[date | None, list[dict], int]:
    """Parse puro do layout oficial do Extrato Consolidado."""
    report_date = _report_date(rows)
    positions: list[dict] = []
    skipped = 0
    for row in rows:
        title = row[0] if row else None
        if not _norm(title).startswith("TESOURO"):
            if _norm(title) == "TOTAL":
                skipped += 1
            continue
        maturity = row[1] if len(row) > 1 else None
        invested = _number(row[2] if len(row) > 2 else None)
        liquid_value = _number(row[4] if len(row) > 4 else None)
        quantity = _number(row[5] if len(row) > 5 else None)
        if not maturity or invested is None or invested < 0 or liquid_value is None or liquid_value < 0 or quantity is None or quantity <= 0:
            skipped += 1
            continue
        positions.append({
            "ticker": tesouro_security_key(title, maturity),
            "name": str(title).strip(),
            "asset_type": "fixed_income",
            "quantity": quantity,
            "market_price": None,
            "market_value": liquid_value,
            "invested_value": invested,
            "is_loaned": False,
            "currency": "BRL",
        })
    return report_date, positions, skipped


def filter_redundant_against_b3(
    positions: list[dict], b3_tickers: Iterable[str],
) -> tuple[list[dict], int]:
    """Remove posicoes ja cobertas por operacoes B3 do mesmo titulo."""
    b3_keys = {tesouro_security_key(ticker, ticker) for ticker in b3_tickers}
    kept = [position for position in positions if position["ticker"] not in b3_keys]
    return kept, len(positions) - len(kept)


def _b3_tickers(conn: Connection, user_id: str) -> list[str]:
    rows = conn.execute(text("""
        SELECT DISTINCT a.ticker
        FROM investment_transactions it
        JOIN assets a ON a.id = it.asset_id
        WHERE it.user_id = :uid
          AND it.external_id LIKE 'b3%'
          AND a.class = 'fixed_income'
    """), {"uid": user_id}).fetchall()
    return [str(row[0]) for row in rows]


def _snapshot_exists(
    conn: Connection, portfolio_id: str, asset_id: str, report_date: date,
) -> bool:
    return conn.execute(text("""
        SELECT 1
        FROM portfolio_position_snapshots
        WHERE portfolio_id = :pid AND asset_id = :aid AND report_date = :report_date
          AND source_system = 'app4'
          AND (
              source_table = 'tesouro_direto'
              OR source_id LIKE 'td-snap-%'
          )
        LIMIT 1
    """), {"pid": portfolio_id, "aid": asset_id, "report_date": report_date}).fetchone() is not None


def parse(file_bytes: bytes, engine: Engine) -> dict[str, Any]:
    """Importa um Extrato Consolidado mensal do Tesouro Direto."""
    summary = make_summary(SOURCE)
    user_id = settings.OWNER_USER_ID
    if not user_id:
        summary["errors"].append("OWNER_USER_ID nao configurado.")
        return finalize_summary(summary)
    if not isinstance(file_bytes, (bytes, bytearray)) or not file_bytes:
        summary["errors"].append("Arquivo vazio ou invalido.")
        return finalize_summary(summary)
    if len(file_bytes) > MAX_FILE_BYTES:
        summary["errors"].append("Arquivo excede o limite de 5 MB.")
        return finalize_summary(summary)

    try:
        workbook = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
        if "Extrato" not in workbook.sheetnames:
            raise ValueError("aba 'Extrato' nao encontrada")
        rows = list(workbook["Extrato"].iter_rows(values_only=True))
        report_date, positions, skipped = _parse_rows(rows)
    except Exception as exc:  # noqa: BLE001
        summary["errors"].append(f"Arquivo invalido: {safe_error(exc)}")
        return finalize_summary(summary)

    summary["rows_skipped"] += skipped
    if report_date is None:
        summary["errors"].append("Periodo MM/AAAA nao encontrado no extrato.")
        return finalize_summary(summary)
    if not positions:
        summary["errors"].append("Nenhuma posicao valida do Tesouro encontrada.")
        return finalize_summary(summary)

    ensure_external_id_columns(engine)
    try:
        with engine.connect() as conn, conn.begin():
            positions, b3_duplicates = filter_redundant_against_b3(
                positions, _b3_tickers(conn, user_id)
            )
            summary["duplicates_skipped"] += b3_duplicates
            portfolio_id = _ensure_portfolio(conn, user_id)
            for position in positions:
                asset_id = get_or_create_asset(
                    conn, position["ticker"], position["name"], "fixed_income", "BRL"
                )
                if _snapshot_exists(conn, portfolio_id, asset_id, report_date):
                    summary["duplicates_skipped"] += 1
                    continue
                source_id = make_external_id("td-snap", [
                    report_date.isoformat(), position["ticker"],
                    f"{position['quantity']:.8f}", f"{position['invested_value']:.2f}",
                    f"{position['market_value']:.2f}",
                ])
                _insert_snapshot(
                    conn, user_id, portfolio_id, asset_id, position, report_date,
                    "Tesouro Direto", source_id,
                    source_table="tesouro_direto",
                )
                summary["positions_imported"] += 1
    except Exception as exc:  # noqa: BLE001
        summary["errors"].append(f"Falha ao gravar extrato: {safe_error(exc)}")

    summary["_report_date"] = report_date.isoformat()
    summary["_institution"] = "Tesouro Direto"
    return finalize_summary(summary)
