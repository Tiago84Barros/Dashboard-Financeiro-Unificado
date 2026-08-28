"""
data_pipeline/importers/investments/xp_consolidado.py
=====================================================
Parser do "Relatório Consolidado" da XP Investimentos (.xlsx).

Suporta tanto o relatório mensal quanto o anual:
  relatorio-consolidado-mensal-2026-janeiro.xlsx → data = último dia do mês
  relatorio-consolidado-anual-2025.xlsx           → data = 31/12/{ano}

Grava em `portfolio_position_snapshots` (não em `investment_transactions`):
o relatório XP é uma FOTO da carteira em uma data, não uma transação. Cada
re-import com a mesma `report_date` substitui o snapshot anterior daquele
arquivo (via UNIQUE em (portfolio_id, asset_id, report_date, source_system,
source_table, source_id)).

Abas processadas:
  Posição - Ações          → asset_type='stock'
  Posição - Empréstimos    → asset_type='stock' + is_loaned=True
  Posição - ETF            → asset_type='etf'
  Posição - Fundos         → 'fii' ou 'fundo_rf' (depende do Tipo)
  Posição - Renda Fixa     → 'renda_fixa' ou 'fundo_rf' (CFF)
  Posição - Tesouro Direto → 'tesouro'

A aba "Proventos Recebidos", quando presente, alimenta `dividends`
(deduplica contra B3 Movimentacao por ativo, data, tipo e valor).
"""
from __future__ import annotations

import calendar
import io
import logging
import re
import unicodedata
from datetime import date
from typing import Any

from sqlalchemy.engine import Connection, Engine

from core.config import settings

from .common import (
    ensure_external_id_columns,
    finalize_summary,
    get_or_create_account,
    get_or_create_asset,
    get_or_create_institution,
    insert_dividend,
    make_external_id,
    make_summary,
    safe_error,
)
from .positions import PORTFOLIO_NAME, PORTFOLIO_TYPE

logger = logging.getLogger(__name__)

SOURCE = "xp_consolidado"
XP_INSTITUTION_NAME = "XP Investimentos S.A."
XP_INSTITUTION_TYPE = "broker"
XP_ACCOUNT_NAME = "XP - Proventos"
XP_ACCOUNT_TYPE = "investment"

_MONTHS_PT: dict[str, int] = {
    "janeiro": 1, "fevereiro": 2, "marco": 3, "marc": 3, "marco_": 3,
    "marc%c3%a7o": 3, "marco%": 3, "abril": 4, "maio": 5, "junho": 6,
    "julho": 7, "agosto": 8, "setembro": 9, "outubro": 10,
    "novembro": 11, "dezembro": 12,
}

# Tipos de fundo que entram como 'fundo_rf' (renda fixa via fundo)
_FUNDO_RF_KEYWORDS = ("renda fixa", "fic rf", "ficrf", "firf")

# Prefixos genéricos de produto que não viram ticker no banco
_GENERIC_TICKER_PREFIXES = {"CDB", "LCI", "LCA", "CFF", "CRI", "CRA",
                             "LF", "LC", "DEB", "CCB"}

# Mapeamento Tipo de Evento (XP) → dividends.type
_XP_INCOME_TYPE_MAP: dict[str, str] = {
    "juros sobre capital próprio":   "jcp",
    "juros sobre capital proprio":   "jcp",
    "jcp":                            "jcp",
    "rendimento":                     "reit_income",
    "dividendo":                      "dividend",
    "dividendo - ação":               "dividend",
    "dividendo - acao":               "dividend",
    "reembolso":                      "reit_income",
    "pagamento de prêmio/rendimentos": "reit_income",
    "pagamento de premio/rendimentos": "reit_income",
    "pagamento de rendimentos":       "reit_income",
    "amortização":                    "amortization",
    "amortizacao":                    "amortization",
}


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _norm(s: str) -> str:
    return unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode().lower()


def _decimal(val: Any) -> float | None:
    if val is None or val == "-" or val == "":
        return None
    try:
        return float(str(val).replace(",", "."))
    except (ValueError, TypeError):
        return None


def _parse_report_date(filename: str) -> date:
    """Infere a data de referência do nome do arquivo XP."""
    name = filename.lower().replace(".xlsx", "")

    m = re.search(r"mensal-(\d{4})-([a-záéíóúç]+)", name)
    if m:
        year = int(m.group(1))
        month = _MONTHS_PT.get(_norm(m.group(2)), 0)
        if month:
            last_day = calendar.monthrange(year, month)[1]
            return date(year, month, last_day)

    m = re.search(r"anual-(\d{4})", name)
    if m:
        return date(int(m.group(1)), 12, 31)

    return date.today()


def _extract_ticker_name(produto: str) -> tuple[str, str]:
    if not produto or not isinstance(produto, str):
        return ("?", "")
    s = produto.strip()
    if " - " in s:
        parts = s.split(" - ", 1)
        return (parts[0].strip().upper(), parts[1].strip())
    return (s.strip().upper(), s.strip())


def _tesouro_ticker(product_name: str, maturity_str: Any) -> tuple[str, str]:
    name = str(product_name or "").strip()
    upper = name.upper()
    year_suffix = ""
    if maturity_str:
        m = re.search(r"(\d{4})", str(maturity_str))
        if m:
            year_suffix = m.group(1)
    if "SELIC" in upper:
        return f"TSELIC{year_suffix}", name
    if "IPCA" in upper and "EDUCA" not in upper and "RENDA" not in upper:
        return f"TIPCA{year_suffix}", name
    if "PREFIXADO" in upper or "PRÉ" in upper or "PRE " in upper:
        return f"TPRE{year_suffix}", name
    if "EDUCA" in upper:
        return f"TEDUCA{year_suffix}", name
    if "RENDA" in upper:
        return f"TRENDA{year_suffix}", name
    return f"TESOUR{year_suffix or 'X'}", name


def _is_fundo_rf(tipo_str: str) -> bool:
    t = (tipo_str or "").strip().lower()
    return any(kw in t for kw in _FUNDO_RF_KEYWORDS)


def _rows(ws, skip_header: int = 1):
    for row in ws.iter_rows(min_row=1 + skip_header, values_only=True):
        if not row or row[0] is None or str(row[0]).strip() == "":
            continue
        yield row


# ─────────────────────────────────────────────────────────────────────────────
# Parsers por aba — devolvem lista de dicts (snapshot rows)
# ─────────────────────────────────────────────────────────────────────────────

def _parse_acoes(ws) -> list[dict]:
    out = []
    for row in _rows(ws):
        ticker, name = _extract_ticker_name(row[0])
        qty = _decimal(row[8]) if len(row) > 8 else None
        price = _decimal(row[12]) if len(row) > 12 else None
        value = _decimal(row[13]) if len(row) > 13 else None
        if qty is None or qty <= 0:
            continue
        out.append({
            "ticker": ticker, "name": name, "asset_type": "stock",
            "quantity": qty, "market_price": price, "market_value": value,
            "invested_value": None, "is_loaned": False, "currency": "BRL",
        })
    return out


def _parse_emprestimos(ws) -> list[dict]:
    out = []
    for row in _rows(ws):
        ticker, name = _extract_ticker_name(row[0])
        qty = _decimal(row[11]) if len(row) > 11 else None
        price = _decimal(row[12]) if len(row) > 12 else None
        value = _decimal(row[13]) if len(row) > 13 else None
        if qty is None or qty <= 0:
            continue
        out.append({
            "ticker": ticker, "name": name, "asset_type": "stock",
            "quantity": qty, "market_price": price, "market_value": value,
            "invested_value": None, "is_loaned": True, "currency": "BRL",
        })
    return out


def _parse_etf(ws) -> list[dict]:
    out = []
    for row in _rows(ws):
        ticker, name = _extract_ticker_name(row[0])
        qty = _decimal(row[7]) if len(row) > 7 else None
        price = _decimal(row[11]) if len(row) > 11 else None
        value = _decimal(row[12]) if len(row) > 12 else None
        if qty is None or qty <= 0:
            continue
        out.append({
            "ticker": ticker, "name": name, "asset_type": "etf",
            "quantity": qty, "market_price": price, "market_value": value,
            "invested_value": None, "is_loaned": False, "currency": "BRL",
        })
    return out


def _parse_fundos(ws) -> list[dict]:
    out = []
    for row in _rows(ws):
        ticker, name = _extract_ticker_name(row[0])
        tipo_col = str(row[6]).strip() if len(row) > 6 and row[6] else ""
        qty = _decimal(row[8]) if len(row) > 8 else None
        price = _decimal(row[12]) if len(row) > 12 else None
        value = _decimal(row[13]) if len(row) > 13 else None
        if qty is None or qty <= 0:
            continue
        asset_type = "fundo_rf" if _is_fundo_rf(tipo_col) else "fii"
        out.append({
            "ticker": ticker, "name": name, "asset_type": asset_type,
            "quantity": qty, "market_price": price, "market_value": value,
            "invested_value": None, "is_loaned": False, "currency": "BRL",
        })
    return out


def _parse_renda_fixa(ws) -> list[dict]:
    out = []
    for row in _rows(ws):
        produto = str(row[0]).strip() if row[0] else ""
        if not produto:
            continue
        tipo = produto.split(" - ")[0].strip() if " - " in produto else produto[:10]
        emissor = str(row[2]).strip() if len(row) > 2 and row[2] else ""
        codigo = str(row[3]).strip() if len(row) > 3 and row[3] else ""
        ticker = codigo.upper() if codigo and codigo != "-" else tipo.upper()
        name = f"{tipo} — {emissor}" if emissor else tipo
        qty = _decimal(row[8]) if len(row) > 8 else None
        invested = _decimal(row[16]) if len(row) > 16 else None  # ValorCurva
        if qty is None or qty <= 0:
            continue
        if invested is None or invested <= 0:
            invested = _decimal(row[14]) if len(row) > 14 else None  # MTM
        atype = "fundo_rf" if tipo.upper().startswith("CFF") else "fixed_income"
        out.append({
            "ticker": ticker, "name": name, "asset_type": atype,
            "quantity": qty, "market_price": None, "market_value": invested,
            "invested_value": invested, "is_loaned": False, "currency": "BRL",
        })
    return out


def _parse_tesouro(ws) -> list[dict]:
    out = []
    for row in _rows(ws):
        produto = row[0]
        if not produto or not isinstance(produto, str):
            continue
        maturity = row[4] if len(row) > 4 else None
        ticker, name = _tesouro_ticker(produto, maturity)
        qty = _decimal(row[5]) if len(row) > 5 else None
        invested = _decimal(row[9]) if len(row) > 9 else None
        market = _decimal(row[12]) if len(row) > 12 else None
        if qty is None or qty <= 0:
            continue
        out.append({
            "ticker": ticker, "name": name, "asset_type": "fixed_income",
            "quantity": qty, "market_price": None,
            "market_value": market or invested,
            "invested_value": invested,
            "is_loaned": False, "currency": "BRL",
        })
    return out


# Tabela de parsers por nome normalizado de aba
_SHEET_PARSERS: list[tuple[str, Any]] = [
    ("posicao - acoes",          _parse_acoes),
    ("posicao - emprestimos",    _parse_emprestimos),
    ("posicao - etf",            _parse_etf),
    ("posicao - fundos",         _parse_fundos),
    ("posicao - renda fixa",     _parse_renda_fixa),
    ("posicao - tesouro direto", _parse_tesouro),
]


# ─────────────────────────────────────────────────────────────────────────────
# Aba "Proventos Recebidos" → dividends
# ─────────────────────────────────────────────────────────────────────────────

def _parse_proventos(
    ws,
    conn: Connection,
    user_id: str,
    summary: dict,
    report_date: date,
) -> None:
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        if not row or row[0] is None:
            continue
        produto_raw = str(row[0]).strip()
        if not produto_raw or produto_raw.lower() in ("produto", "total", "-"):
            continue
        try:
            pagamento_raw = row[1] if len(row) > 1 else None
            tipo_evento = str(row[2]).strip() if len(row) > 2 and row[2] else ""
            valor_raw = row[6] if len(row) > 6 else None

            if not pagamento_raw or not tipo_evento or valor_raw is None:
                summary["rows_skipped"] += 1
                continue

            ticker = (
                produto_raw.split(" - ")[0].strip().upper()
                if " - " in produto_raw else produto_raw[:10].strip().upper()
            )
            if ticker in _GENERIC_TICKER_PREFIXES:
                summary["rows_skipped"] += 1
                continue

            inc_type = _XP_INCOME_TYPE_MAP.get(_norm(tipo_evento))
            if inc_type is None:
                summary["rows_skipped"] += 1
                continue

            amount = _decimal(valor_raw)
            if amount is None or amount <= 0:
                summary["rows_skipped"] += 1
                continue

            from datetime import datetime as _dt
            if isinstance(pagamento_raw, _dt):
                pay_date = pagamento_raw.date()
            elif isinstance(pagamento_raw, date):
                pay_date = pagamento_raw
            else:
                try:
                    pay_date = _dt.strptime(
                        str(pagamento_raw).strip(), "%d/%m/%Y"
                    ).date()
                except ValueError:
                    summary["rows_skipped"] += 1
                    continue

            # Asset deve existir — XP relata produtos que talvez nao estejam
            # no nosso assets ainda. Cria como stock por padrao (XP nao
            # informa classe na aba Proventos).
            asset_id = get_or_create_asset(
                conn, ticker=ticker, name=ticker, asset_class="stock",
            )

            ext_id = make_external_id(
                "xpcsl-inc",
                [pay_date.isoformat(), ticker, inc_type, f"{amount:.2f}"],
            )

            try:
                with conn.begin_nested():
                    new_id = insert_dividend(
                        conn,
                        user_id=user_id,
                        asset_id=asset_id,
                        div_type=inc_type,
                        amount_per_unit=amount,  # XP nao separa por cota
                        quantity=1.0,
                        total_amount=amount,
                        ex_date=None,
                        payment_date=pay_date,
                        external_id=ext_id,
                    )
                    if new_id is None:
                        summary["duplicates_skipped"] += 1
                    else:
                        summary["incomes_imported"] += 1
            except Exception as exc:  # noqa: BLE001
                summary["errors"].append(
                    f"Proventos linha {i + 2}: {safe_error(exc)}"
                )

        except Exception as exc:  # noqa: BLE001
            summary["errors"].append(
                f"Proventos linha {i + 2}: {safe_error(exc)}"
            )


# ─────────────────────────────────────────────────────────────────────────────
# Setup de portfolio + persistência de snapshots
# ─────────────────────────────────────────────────────────────────────────────

def _ensure_portfolio(conn: Connection, user_id: str) -> str:
    row = conn.execute(
        __import__("sqlalchemy").text("""
            SELECT id FROM portfolios
            WHERE user_id = :uid AND name = :name
            LIMIT 1
        """),
        {"uid": user_id, "name": PORTFOLIO_NAME},
    ).fetchone()
    if row:
        return str(row[0])

    row = conn.execute(
        __import__("sqlalchemy").text("""
            INSERT INTO portfolios (user_id, name, type, active)
            VALUES (:uid, :name, :type, TRUE)
            RETURNING id
        """),
        {"uid": user_id, "name": PORTFOLIO_NAME, "type": PORTFOLIO_TYPE},
    ).fetchone()
    return str(row[0])


def _ensure_xp_account(conn: Connection, user_id: str) -> str:
    inst_id = get_or_create_institution(
        conn, XP_INSTITUTION_NAME, XP_INSTITUTION_TYPE,
    )
    return get_or_create_account(
        conn,
        user_id=user_id,
        institution_id=inst_id,
        name=XP_ACCOUNT_NAME,
        account_type=XP_ACCOUNT_TYPE,
        currency="BRL",
    )


def _insert_snapshot(
    conn: Connection,
    user_id: str,
    portfolio_id: str,
    asset_id: str,
    pos: dict,
    report_date: date,
    institution: str,
    source_id: str,
    *,
    source_table: str = "xp_consolidado",
) -> str | None:
    """Insere uma posição em portfolio_position_snapshots (idempotente)."""
    from sqlalchemy import text
    row = conn.execute(
        text("""
            INSERT INTO portfolio_position_snapshots
                (user_id, portfolio_id, asset_id, report_date,
                 quantity, market_price, market_value, invested_value,
                 asset_name, asset_type, is_loaned, institution,
                 currency, country, source_system, source_table, source_id)
            VALUES
                (:uid, :pid, :aid, :rd,
                 :qty, :mp, :mv, :iv,
                 :an, :at, :il, :inst,
                 :ccy, :country, 'app4', :source_table, :sid)
            ON CONFLICT (portfolio_id, asset_id, report_date,
                         source_system, source_table, source_id)
            DO UPDATE SET
                quantity       = EXCLUDED.quantity,
                market_price   = EXCLUDED.market_price,
                market_value   = EXCLUDED.market_value,
                invested_value = EXCLUDED.invested_value,
                asset_name     = EXCLUDED.asset_name,
                asset_type     = EXCLUDED.asset_type,
                is_loaned      = EXCLUDED.is_loaned,
                institution    = EXCLUDED.institution,
                imported_at    = NOW()
            RETURNING id
        """),
        {
            "uid":     user_id,
            "pid":     portfolio_id,
            "aid":     asset_id,
            "rd":      report_date,
            "qty":     pos["quantity"],
            "mp":      pos.get("market_price"),
            "mv":      pos.get("market_value") or 0,
            "iv":      pos.get("invested_value"),
            "an":      pos["name"][:300] if pos.get("name") else None,
            "at":      pos["asset_type"],
            "il":      bool(pos.get("is_loaned")),
            "inst":    institution[:150] if institution else None,
            "ccy":     pos.get("currency", "BRL"),
            "country": "BR",
            "source_table": source_table,
            "sid":     source_id,
        },
    ).fetchone()
    return str(row[0]) if row else None


# ─────────────────────────────────────────────────────────────────────────────
# Entry point
# ─────────────────────────────────────────────────────────────────────────────

def parse(
    file_input: bytes | tuple[str, bytes] | list[tuple[str, bytes]],
    engine: Engine,
) -> dict[str, Any]:
    """
    Processa um OU vários relatórios consolidados da XP (multi-arquivo).

    `file_input` pode ser:
      - bytes simples (filename inferido como 'xp.xlsx')
      - (filename, bytes)
      - list[(filename, bytes)] — caso "múltiplos arquivos de uma vez".

    Quando recebe lista, processa cada arquivo individualmente e agrega
    contadores num único summary. Erros em um arquivo não abortam os
    outros — ficam registrados em summary["errors"].
    """
    # Normaliza entrada para lista de (filename, bytes)
    if isinstance(file_input, (bytes, bytearray)):
        items = [("xp.xlsx", bytes(file_input))]
    elif isinstance(file_input, tuple) and len(file_input) == 2:
        items = [(str(file_input[0]), bytes(file_input[1]))]
    elif isinstance(file_input, list):
        items = [(str(name), bytes(data)) for name, data in file_input]
    else:
        summary = make_summary(SOURCE)
        summary["status"] = "failed"
        summary["errors"].append(
            f"Tipo de entrada invalido: {type(file_input).__name__}"
        )
        return finalize_summary(summary)

    if not items:
        summary = make_summary(SOURCE)
        summary["errors"].append("Nenhum arquivo recebido.")
        return finalize_summary(summary)

    # Caso single: delega direto (preserva semântica original)
    if len(items) == 1:
        return _parse_single(items[0], engine)

    # Caso multi: itera e agrega contadores
    agg = make_summary(SOURCE)
    files_summaries: list[str] = []
    for item in items:
        try:
            s = _parse_single(item, engine)
        except Exception as exc:  # noqa: BLE001
            agg["errors"].append(f"{item[0]}: {safe_error(exc)}")
            files_summaries.append(f"{item[0]}: FAILED")
            continue
        for key in (
            "transactions_imported", "incomes_imported", "positions_imported",
            "duplicates_skipped", "rows_skipped", "files_skipped",
        ):
            agg[key] = int(agg.get(key, 0)) + int(s.get(key, 0))
        if s.get("errors"):
            for e in s["errors"]:
                agg["errors"].append(f"{item[0]}: {e}")
        if s.get("files_skipped_notes"):
            agg.setdefault("files_skipped_notes", []).extend(
                s["files_skipped_notes"]
            )
        files_summaries.append(
            f"{item[0]}: {s.get('status', '?')} "
            f"(tx={s.get('transactions_imported', 0)}, "
            f"inc={s.get('incomes_imported', 0)}, "
            f"pos={s.get('positions_imported', 0)})"
        )

    # Anota no campo de notas pra UI mostrar resumo por arquivo
    agg.setdefault("files_skipped_notes", []).insert(
        0, f"Processados {len(items)} arquivos: " + " | ".join(files_summaries)
    )
    return finalize_summary(agg)


def _parse_single(
    file_input: tuple[str, bytes],
    engine: Engine,
) -> dict[str, Any]:
    """
    Processa um único relatório consolidado da XP.

    `file_input` é (filename, bytes). A data do snapshot é inferida
    do nome do arquivo.
    """
    summary = make_summary(SOURCE)
    user_id = settings.OWNER_USER_ID
    if not user_id:
        summary["status"] = "failed"
        summary["errors"].append("OWNER_USER_ID nao configurado.")
        return finalize_summary(summary)

    filename, file_bytes = file_input
    file_bytes = bytes(file_bytes)

    ensure_external_id_columns(engine)

    try:
        import openpyxl
    except ImportError:
        summary["status"] = "failed"
        summary["errors"].append("openpyxl nao instalado.")
        return finalize_summary(summary)

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as exc:  # noqa: BLE001
        summary["status"] = "failed"
        summary["errors"].append(f"Arquivo invalido: {safe_error(exc)}")
        return finalize_summary(summary)

    report_date = _parse_report_date(filename)

    # Detecta nome da instituição na 2ª linha da 1ª aba (padrão XP)
    institution = "XP Investimentos"
    for sheet in wb.worksheets:
        for row in sheet.iter_rows(min_row=2, max_row=2, values_only=True):
            if row and len(row) > 1 and row[1] and isinstance(row[1], str):
                institution = row[1].strip()[:150]
                break
        break

    sheet_map = {_norm(ws.title): ws for ws in wb.worksheets}

    with engine.connect() as conn:
        with conn.begin():
            try:
                portfolio_id = _ensure_portfolio(conn, user_id)
            except Exception as exc:  # noqa: BLE001
                summary["status"] = "failed"
                summary["errors"].append(
                    f"Falha ao preparar carteira XP: {safe_error(exc)}"
                )
                return finalize_summary(summary)

            # Processa cada aba de posição
            for sheet_key, parser_fn in _SHEET_PARSERS:
                ws = sheet_map.get(sheet_key)
                if not ws:
                    continue
                try:
                    positions = parser_fn(ws)
                except Exception as exc:  # noqa: BLE001
                    summary["errors"].append(
                        f"Aba '{ws.title}': {safe_error(exc)}"
                    )
                    continue

                for pos in positions:
                    try:
                        with conn.begin_nested():
                            asset_id = get_or_create_asset(
                                conn,
                                ticker=pos["ticker"],
                                name=pos.get("name") or pos["ticker"],
                                asset_class=_map_asset_type(pos["asset_type"]),
                                currency=pos.get("currency", "BRL"),
                            )
                            source_id = make_external_id(
                                "xpcsl-snap",
                                [report_date.isoformat(),
                                 pos["ticker"],
                                 pos["asset_type"],
                                 f"{pos['quantity']:.6f}"],
                            )
                            _insert_snapshot(
                                conn,
                                user_id=user_id,
                                portfolio_id=portfolio_id,
                                asset_id=asset_id,
                                pos=pos,
                                report_date=report_date,
                                institution=institution,
                                source_id=source_id,
                            )
                            summary["positions_imported"] += 1
                    except Exception as exc:  # noqa: BLE001
                        summary["errors"].append(
                            f"[{ws.title}] {pos.get('ticker', '?')}: "
                            f"{safe_error(exc)}"
                        )

            # Aba Proventos Recebidos → dividends. A camada comum de insert
            # deduplica contra B3 Movimentação pela chave canônica, além do
            # external_id específico da XP.
            ws_prov = sheet_map.get("proventos recebidos")
            if ws_prov:
                try:
                    _ensure_xp_account(conn, user_id)
                except Exception as exc:  # noqa: BLE001
                    summary["errors"].append(
                        f"Falha ao preparar conta XP: {safe_error(exc)}"
                    )
                else:
                    _parse_proventos(
                        ws_prov, conn, user_id, summary, report_date,
                    )

    summary["_report_date"] = report_date.isoformat()
    summary["_institution"] = institution
    return finalize_summary(summary)


def _map_asset_type(xp_type: str) -> str:
    """Converte asset_type do snapshot XP para o CHECK de assets.class."""
    return {
        "stock":        "stock",
        "etf":          "etf",
        "fii":          "reit",
        "fundo_rf":     "fixed_income",
        "renda_fixa":   "fixed_income",
        "fixed_income": "fixed_income",
        "tesouro":      "fixed_income",
    }.get(xp_type, "other")
